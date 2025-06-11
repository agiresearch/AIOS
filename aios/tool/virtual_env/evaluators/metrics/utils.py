import builtins
import datetime
import functools
import itertools
import logging
import operator
import re
import zipfile
import pandas as pd
from typing import Any, TypeVar, Union, Iterable, Optional, Callable
from typing import Dict, List, Set, Match, Tuple, Pattern
from urllib.parse import urlparse, urlunparse

import formulas
import lxml.cssselect
import lxml.etree
import xmltodict
from lxml.etree import _Element
from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.chart._chart import ChartBase
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.pivot.cache import CacheSource as PivotCacheSource
from openpyxl.pivot.table import TableDefinition as PivotTableDefinition
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import coordinate_to_tuple, get_column_letter
from openpyxl.worksheet.cell_range import MultiCellRange, CellRange
from openpyxl.worksheet.dimensions import DimensionHolder
from openpyxl.worksheet.filters import AutoFilter, SortState
from openpyxl.worksheet.worksheet import Worksheet

V = TypeVar("Value")

logger = logging.getLogger("desktopenv.metrics.utils")

_xlsx_namespaces = [("oo", "http://schemas.openxmlformats.org/spreadsheetml/2006/main")
    , ("x14", "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main")
    , ("xm", "http://schemas.microsoft.com/office/excel/2006/main")
                    ]
_xlsx_ns_mapping = dict(_xlsx_namespaces)
_xlsx_ns_imapping = dict(map(lambda itm: (itm[1], itm[0]), _xlsx_namespaces))
_xlsx_ns_imapping["http://schemas.openxmlformats.org/spreadsheetml/2006/main"] = None
_sheet_name_selector = lxml.cssselect.CSSSelector("oo|sheets>oo|sheet", namespaces=_xlsx_ns_mapping)
_sparklines_selector = lxml.cssselect.CSSSelector("x14|sparkline", namespaces=_xlsx_ns_mapping)


def load_sparklines(xlsx_file: str, sheet_name: str) -> Dict[str, str]:
    #  function load_sparklines {{{ # 
    """
    Args:
        xlsx_file (str): path to xlsx
        sheet_name (str): sheet name

    Returns:
        List[Dict[str, str]]: sparkline definitions in form of
          {
            "F3": "Sheet1!C3:E3"
          }
    """

    # read xlsx
    try:
        with zipfile.ZipFile(xlsx_file, "r") as z_f:
            with z_f.open("xl/workbook.xml") as f:
                workbook_database: _Element = lxml.etree.fromstring(f.read())
                sheets: List[_Element] = _sheet_name_selector(workbook_database)
                sheet_names: Dict[str, str] = {sh.get("name"): sh.get("sheetId") for sh in sheets}
            with z_f.open("xl/worksheets/sheet{:}.xml".format(sheet_names[sheet_name])) as f:
                sheet: _Element = lxml.etree.fromstring(f.read())
                sparklines: List[_Element] = _sparklines_selector(sheet)
    except zipfile.BadZipFile:
        return {}

    sparklines_dict: Dict[str, str] = {}
    for sp_l in sparklines:
        sparkline_xml: str = lxml.etree.tostring(sp_l, encoding="unicode")
        sparkline: Dict[str, Dict[str, str]] = xmltodict.parse(sparkline_xml
                                                               , process_namespaces=True
                                                               , namespaces=_xlsx_ns_imapping
                                                               )
        sparklines_dict[sparkline["x14:sparkline"]["xm:sqref"]] = sparkline["x14:sparkline"]["xm:f"]
    return sparklines_dict
    #  }}} function load_sparklines # 


# Available Chart Properties:
# title: str
# anchor: ["oneCell" | "twoCell" | "absolute", col0, row0, col1, row1]
# legend: "b" | "tr" | "l" | "r" | "t"
# width: number
# height: number
# type: "scatterChart" | "lineChart" | "barChart"
# direction: "bar" (hori) | "col" (vert)
# xtitle, ytitle, ztitle: str
def load_charts(xlsx_file: Workbook, sheet_name: str, **options) -> Dict[str, Any]:
    #  function load_charts {{{ # 
    """
    Args:
        xlsx_file (Workbook): concerned excel book
        sheet_name (str): sheet name
        options (Dict[str, List[str]]): dict like {"chart_props": list of str}
          giving the concerned chart properties

    Returns:
        Dict[str, Any]: information of charts, dict like
          {
            <str representing data source>: {
                <str as property>: anything
            }
          }
    """

    # workbook: Workbook = openpyxl.load_workbook(filename=xlsx_file)
    try:
        worksheet: Worksheet = xlsx_file[sheet_name]
    except KeyError:
        return {}
    charts: List[ChartBase] = worksheet._charts

    chart_set: Dict[str, Any] = {}
    chart_props: Set[str] = set(options["chart_props"]) if "chart_props" in options else set()
    for ch in charts:
        series: List[str] = []
        for ser in ch.series:
            if hasattr(ser.val, "numRef") and hasattr(ser.val.numRef, "f"):
                value_str: str = ser.val.numRef.f
            elif hasattr(ser.val, "strRef") and hasattr(ser.val.strRef, "f"):
                value_str: str = ser.val.strRef.f
            else:
                value_str: str = ""
            if hasattr(ser.cat, "numRef") and hasattr(ser.cat.numRef, "f"):
                categ_str: str = ser.cat.numRef.f
            elif hasattr(ser.cat, "strRef") and hasattr(ser.cat.strRef, "f"):
                categ_str: str = ser.cat.strRef.f
            else:
                categ_str: str = ""
            series.append("{:},{:}".format(value_str, categ_str))
        series: str = ";".join(series)

        # TODO: maybe more aspects, like chart type
        info: Dict[str, Any] = {}

        if "title" in chart_props:
            try:
                info["title"] = ch.title.tx.rich.p[0].r[0].t
            except:
                info["title"] = None
        if "legend" in chart_props:
            info["legend"] = ch.legend.position if ch.legend is not None else None
        if "anchor" in chart_props:
            info["anchor"] = [ch.anchor.editAs
                , ch.anchor._from.col, ch.anchor.to.row
                , ch.anchor.to.col, ch.anchor.to.row
                              ]
        if "width" in chart_props:
            info["width"] = ch.width
        if "height" in chart_props:
            info["height"] = ch.height
        if "type" in chart_props:
            info["type"] = ch.tagname
        if "direction" in chart_props:
            info["direction"] = ch.barDir

        if "xtitle" in chart_props:
            try:
                info["xtitle"] = ch.x_axis.title.tx.rich.p[0].r[0].t
            except:
                info["xtitle"] = None
        if "ytitle" in chart_props:
            try:
                info["ytitle"] = ch.y_axis.title.tx.rich.p[0].r[0].t
            except:
                info["ytitle"] = None
        if "ztitle" in chart_props:
            try:
                info["ztitle"] = ch.z_axis.title.tx.rich.p[0].r[0].t
            except:
                info["ztitle"] = None
        chart_set[series] = info
    logger.debug(".[%s].charts: %s", sheet_name, repr(chart_set))
    return chart_set
    #  }}} function load_charts # 


# Available Pivot Properties:
# name: str
# show_total, show_empty_row, show_empty_col, show_headers: bool
# location: str
# selection: if the concrete item selection should be checked, a list of set of tuple like (bool, index) will be returned; list will be returned instead of set if "ordered" is specified
# filter: if the filter fields should be checked; fields indices will be return in `filter_fields` item
# col_fields: indices
# row_fields: indices
# data_fields: list of str representations. the str representation is like "index;name;subtotal_type;show_data_as"; name is optional and is only returned when `data_fields_name` is specified in `pivot_props`
def load_pivot_tables(xlsx_file: Workbook, sheet_name: str, **options) -> Dict[str, Any]:
    #  function load_pivot_tables {{{ # 
    """
    Args:
        xlsx_file (Workbook): concerned excel book
        sheet_name (str): sheet name
        options (Dict[str, List[str]]): dict like {"pivot_props": list of str}
          giving the concerned pivot properties

    Returns:
        Dict[str, Any]: information of pivot tables, dict like
          {
            <str representing data source>: {
                <str as property>: anything
            }
          }
    """

    try:
        worksheet: Worksheet = xlsx_file[sheet_name]
    except KeyError:
        return {}
    pivots: List[PivotTableDefinition] = worksheet._pivots

    pivot_set: Dict[str, Any] = {}
    pivot_props: Set[str] = set(options.get("pivot_props", []))
    for pvt in pivots:
        raw_selection: List[List[tuple[Optional[bool], int]]] = \
            [[(itm.h, itm.x) for itm in f.items if itm.x is not None] \
             for f in pvt.pivotFields
             ]
        raw__selection: List[List[tuple[Optional[bool], int]]] = list(
            itertools.dropwhile(lambda r: len(r) == 0, raw_selection))
        left_bias = len(raw_selection) - len(raw__selection)
        selection: List[List[tuple[Optional[bool], int]]] = list(
            (itertools.dropwhile(lambda r: len(r) == 0, reversed(raw__selection))))[::-1]
        right_bias = len(raw__selection) - len(selection)
        cache_source: PivotCacheSource = pvt.cache.cacheSource
        cell_range1: str
        cell_range2: str
        cell_range1, cell_range2 = cache_source.worksheetSource.ref.split(":")
        cell_range1: Tuple[int, int] = coordinate_to_tuple(cell_range1)
        cell_range1 = (cell_range1[0], cell_range1[1] + left_bias)
        cell_range2: Tuple[int, int] = coordinate_to_tuple(cell_range2)
        cell_range2 = (cell_range2[0], cell_range2[1] - right_bias)
        source: str = "{:};{:}:{:};{:}".format(cache_source.type, cell_range1, cell_range2,
                                               cache_source.worksheetSource.sheet)

        info: Dict[str, Any] = {}
        if "name" in pivot_props:
            info["name"] = pvt.name

        if "show_total" in pivot_props:
            info["show_total"] = pvt.visualTotals
        if "show_empty_row" in pivot_props:
            info["show_empty_row"] = pvt.showEmptyRow
        if "show_empty_col" in pivot_props:
            info["show_empty_col"] = pvt.showEmptyCol
        if "show_headers" in pivot_props:
            info["show_headers"] = pvt.showHeaders

        if "location" in pivot_props:
            info["location"] = pvt.location
        if "filter" in pivot_props or "selection" in pivot_props:
            info["selection"] = selection if "ordered" in pivot_props else list(set(r) for r in selection)
        if "filter" in pivot_props:
            info["filter_fields"] = set(f.fld for f in pvt.pageFields)
        if "col_fields" in pivot_props:
            info["col_fields"] = [f.x - left_bias for f in pvt.colFields]
        if "row_fields" in pivot_props:
            info["row_fields"] = [f.x - left_bias for f in pvt.rowFields]
        if "data_fields" in pivot_props:
            info["data_fields"] = [
                "{:d};{:};{:};{:}".format(f.fld - left_bias, f.name if "data_fields_name" in pivot_props else ""
                                          , f.subtotal, f.showDataAs
                                          ) \
                for f in pvt.dataFields
                ]

        pivot_set[source] = info
    logger.debug(".[%s].pivots: %s", sheet_name, repr(pivot_set))
    return pivot_set
    #  }}} function load_pivot_tables # 


_shared_str_selector = lxml.cssselect.CSSSelector("oo|sst>oo|si", namespaces=_xlsx_ns_mapping)
_shared_str_value_selector = lxml.cssselect.CSSSelector("oo|t", namespaces=_xlsx_ns_mapping)


def read_cell_value(xlsx_file: str, sheet_name: str, coordinate: str) -> Any:
    #  read_cell_value {{{ # 
    try:
        with zipfile.ZipFile(xlsx_file, "r") as z_f:
            try:
                with z_f.open("xl/sharedStrings.xml") as f:
                    shared_str_xml: _Element = lxml.etree.fromstring(f.read())
                    str_elements: List[_Element] = _shared_str_selector(shared_str_xml)
                    shared_strs: List[str] = [ "".join(t.text for t in _shared_str_value_selector(elm))\
                                           for elm in str_elements
                                             ]
            except:
                logger.debug("Read shared strings error: %s", xlsx_file)

            with z_f.open("xl/workbook.xml") as f:
                workbook_database: _Element = lxml.etree.fromstring(f.read())
                sheets: List[_Element] = _sheet_name_selector(workbook_database)
                sheet_names: Dict[str, str] = {sh.get("name"): sh.get("sheetId") for sh in sheets}

            with z_f.open("xl/worksheets/sheet{:}.xml".format(sheet_names[sheet_name])) as f:
                sheet: _Element = lxml.etree.fromstring(f.read())
                cells: List[_Element] = \
                    lxml.cssselect.CSSSelector('oo|row>oo|c[r="{:}"]'.format(coordinate)
                                               , namespaces=_xlsx_ns_mapping
                                               )(sheet)
                if len(cells) == 0:
                    return None
                cell: _Element = cells[0]
    except zipfile.BadZipFile:
        return None

    cell: Dict[str, str] = xmltodict.parse(lxml.etree.tostring(cell, encoding="unicode")
                                           , process_namespaces=True
                                           , namespaces=_xlsx_ns_imapping
                                           )
    logger.debug("%s.%s[%s]: %s", xlsx_file, sheet_name, coordinate, repr(cell))
    try:
        if "@t" not in cell["c"] or cell["c"]["@t"] == "n":
            return float(cell["c"]["v"])
        if cell["c"]["@t"] == "s":
            return shared_strs[int(cell["c"]["v"])]
        if cell["c"]["@t"] == "str":
            return cell["c"]["v"]
    except (KeyError, ValueError):
        return None
    #  }}} read_cell_value # 


# Supported Styles:
# number_format
# font_name - str
# font_family - float
# font_color - in aRGB, e.g., FF000000 is black
# font_bold - bool
# font_italic - bool
# font_underline - "single" | "double" | "singleAccounting" | "doubleAccounting"
# font_size - float
# fill_type - "patternFill" | "gradientFill"
# bgcolor - in aRGB, e.g., FFFF0000 is red
# fgcolor - in aRGB, e.g., FF00FFFF is yellow
# hyperlink - str
def _read_cell_style(style_name: str, cell: Cell, diff_style: Optional[DifferentialStyle] = None) -> Any:
    if style_name == "number_format":
        return (cell.number_format if diff_style is None else diff_style.numFmt.formatCode) \
            if cell.value is not None and cell.data_type == "n" else None
    elif style_name == "font_name":
        return (diff_style or cell).font.name if cell.value is not None else None
    elif style_name == "font_family":
        return (diff_style or cell).font.family if cell.value is not None else None
    elif style_name == "font_color":
        return (diff_style or cell).font.color.rgb if cell.value is not None else None
    elif style_name == "font_bold":
        return (diff_style or cell).font.bold if cell.value is not None else None
    elif style_name == "font_italic":
        return (diff_style or cell).font.italic if cell.value is not None else None
    elif style_name == "font_underline":
        return (diff_style or cell).font.underline if cell.value is not None else None
    elif style_name == "font_size":
        return (diff_style or cell).font.size if cell.value is not None else None
    elif style_name == "fill_type":
        try:
            return (diff_style or cell).fill.tagname
        except:
            return None
    elif style_name == "bgcolor":
        try:
            return (diff_style or cell).fill.bgColor.rgb
        except:
            return None
    elif style_name == "fgcolor":
        try:
            return (diff_style or cell).fill.fgColor.rgb
        except:
            return None
    elif style_name == "hyperlink":
        return cell.hyperlink or "" if cell.value is not None else None
    else:
        raise NotImplementedError("Unsupported Style: {:}".format(style_name))


_absolute_range_pattern: Pattern[str] = re.compile(r"""\$(?P<col1>[A-Z]{1,3})\$(?P<row1>\d+) # coord1
                                                        (?::
                                                          \$(?P<col2>[A-Z]{1,3})\$(?P<row2>\d+) # coord2
                                                        )?
                                                     """
                                                   , re.X
                                                   )


def load_xlsx_styles(xlsx_file: Workbook, sheet_name: str, book_name: str, **options) -> Dict[str, List[Any]]:
    #  function load_xlsx_styles {{{ # 
    """
    Args:
        xlsx_file (Workbook): concerned excel book
        sheet_name (str): sheet name
        book_name (str): book name
        options (Dict[str, List[str]): dick like {"props": list of str} giving
          the concerned styles

    Returns:
        Dict[str, List[Any]]: dict like
          {
            <str as cell coordinates>: list of anything indicating concerned
              property values
          }
    """

    try:
        worksheet: Worksheet = xlsx_file[sheet_name]
    except KeyError:
        return {}

    style_dict: Dict[str, List[Any]] = {}
    concerned_styles: List[str] = options.get("props", [])

    # Handles Cell Styles
    for col in worksheet.iter_cols():
        for c in col:
            style_list: List[Any] = []
            for st in concerned_styles:
                style_list.append(_read_cell_style(st, c))
            style_dict[c.coordinate] = style_list

    # Handles Conditional Formatting
    conditional_formattings: ConditionalFormattingList = worksheet.conditional_formatting
    formula_parser = formulas.Parser()
    for fmt in conditional_formattings:
        for r in fmt.rules:
            active_cells: List[Cell] = []
            if r.type == "expression":
                condition: Callable[[str], bool] = formula_parser.ast("=" + r.formula[0])[1].compile()
                logger.debug("Expression condition: %s", r.formula[0])

                arguments: List[Any] = []
                absolute_range_match: List[Tuple[str, str, str, str]] = _absolute_range_pattern.findall(r.formula[0])
                for m in absolute_range_match:
                    logger.debug("Absolute ranges: %s", repr(m))
                    if m[2] is None and m[3] is None:
                        arguments.append(read_cell_value(book_name, sheet_name, coordinate="{:}{:}".format(m[0], m[1])))
                    else:
                        arguments.append([read_cell_value(book_name, sheet_name
                                                          , coordinate="{:}{:}".format(get_column_letter(c[1])
                                                                                       , c[0]
                                                                                       )
                                                          ) \
                                          for c in CellRange("{:}{:}:{:}{:}".format(m[0], m[1], m[2], m[3])).cells \
                                          ]
                                         )
                logger.debug("Absolute range arguments: %s", repr(arguments))

                for rge in fmt.cells:
                    for c in rge.cells:
                        cell: Cell = worksheet.cell(row=c[0], column=c[1])
                        cell_value = read_cell_value(book_name, sheet_name
                                                     , coordinate="{:}{:d}".format(get_column_letter(c[1])
                                                                                   , c[0]
                                                                                   )
                                                     )
                        if condition(cell_value, *arguments):
                            logger.debug("Active Cell %s(%s) for %s", repr(cell), str(cell_value), r.formula[0])
                            active_cells.append(cell)
            else:
                raise NotImplementedError("Not Implemented Condition Type: {:}".format(r.type))

            for c in active_cells:
                style_dict[c.coordinate] = [_read_cell_style(st, c, r.dxf) for st in concerned_styles]

    logger.debug(".[%s].styles: %s", sheet_name, repr(style_dict))
    return style_dict
    #  }}} function load_xlsx_styles # 


# Available Row Properties:
# hidden
# collapsed
# height
#
# Available Column Properties:
# width
# auto_size
# hidden
# collapsed
# min
# max
def load_rows_or_cols(xlsx_file: Workbook, sheet_name: str, **options) \
        -> Dict[Union[int, str], Dict[str, Any]]:
    #  function load_rows_or_cols {{{ # 
    """
    Args:
        xlsx_file (Workbook): concerned excel book
        sheet_name (str): sheet name
        options (Dict[str, List[str]]): dict like
          {"obj": "row" | "column", "props": list of str} giving the concerned
          row/column properties

    Returns:
        Dict[Union[int, str], Dict[str, Any]]: row/column information
    """

    try:
        worksheet: Worksheet = xlsx_file[sheet_name]
    except KeyError:
        return {}
    objs: DimensionHolder = getattr(worksheet, "{:}_dimensions".format(options["obj"]))

    obj_set: Dict[int, Any] = {}
    obj_props: Set[str] = set(options.get("props", []))
    for obj_no, obj_dms in objs.items():
        info_dict: Dict[str, Any] = {}
        for prop in obj_props:
            info_dict[prop] = getattr(obj_dms, prop)
        obj_set[obj_no] = info_dict
    return obj_set
    #  }}} function load_rows_or_cols # 


def load_filters(xlsx_file: Workbook, sheet_name: str, **options) -> Dict[str, Any]:
    #  function load_filters {{{ # 
    try:
        worksheet: Worksheet = xlsx_file[sheet_name]
    except KeyError:
        return {}

    filters: AutoFilter = worksheet.auto_filter
    filter_dict: Dict[str, Any] = {}
    filter_dict["ref"] = filters.ref

    # filterColumn
    filter_column_set: List[Dict[str, Any]] = []
    for flt_clm in filters.filterColumn:
        filter_column: Dict[str, Any] = {}
        filter_column["col_id"] = flt_clm.colId
        filter_column["hidden_button"] = flt_clm.hiddenButton
        filter_column["show_button"] = flt_clm.showButton
        if flt_clm.filters is not None:
            filter_column["filters_blank"] = flt_clm.filters.blank
            filter_column["filters"] = set(flt_clm.filters.filter)
        if flt_clm.customFilters is not None:
            filter_column["custom_filters_op"] = flt_clm.customFilters._and
            filter_column["custom_filters"] = set((flt.operator
                                                   , flt.val
                                                   ) \
                                                  for flt in flt_clm.customFilters.customFilter
                                                  )
        filter_column_set.append(filter_column)
    filter_column_set = list(sorted(filter_column_set
                                    , key=(lambda d: d["col_id"])
                                    )
                             )
    filter_dict["filter_column"] = filter_column_set

    # sortState
    sort_state: Optional[SortState] = filters.sortState
    if sort_state is not None:
        sort_state_dict: Dict[str, Any] = {}
        sort_state_dict["sort"] = sort_state.columnSort
        sort_state_dict["case"] = sort_state.caseSensitive
        sort_state_dict["method"] = sort_state.sortMethod
        sort_state_dict["ref"] = sort_state.ref
        sort_state_dict["condition"] = list({"descending": cdt.descending
                                                , "key": cdt.sortBy
                                                , "ref": cdt.ref
                                                , "custom_list": cdt.customList
                                                , "dxf_id": cdt.dxfId
                                                , "icon": cdt.iconSet
                                                , "iconid": cdt.iconId
                                             } \
                                            for cdt in sort_state.sortCondition
                                            )
        filter_dict["sort_state"] = sort_state_dict

    return filter_dict
    #  }}} function load_filters # 


def _match_record(pattern: Dict[str, Any], item: Dict[str, Any]) -> bool:
    return all(k in item and item[k] == val for k, val in pattern.items())


def _multicellrange_containsby(subset_candidate: MultiCellRange, superset_candidate: MultiCellRange) -> bool:
    return all(r in superset_candidate for r in subset_candidate)


def _match_value_to_rule(value: V, rule: Dict[str, Union[str, V]]) -> bool:
    """
    Args:
        value (V): value to match
        rule (Dict[str, Union[str, V]]): rule dict like
          {
            "method": str
            "ref": V as ref value
          }

    Returns:
        bool
    """

    if rule["method"].startswith("re"): # re.FLAGs
        flags: List[str] = rule["method"].split(".")[1:]
        flags: Iterable[re.RegexFlag] = (getattr(re, fl) for fl in flags)
        flag: re.RegexFlag = functools.reduce(operator.or_, flags, re.RegexFlag(0))
        logger.debug("REFLAG: %s", repr(flag))

        match_: Optional[Match[str]] = re.search(rule["ref"], value, flag)
        return match_ is not None
    if rule["method"] in {"eq", "ne"
        , "le", "lt"
        , "ge", "gt"
                          }:
        return getattr(operator, rule["method"])(value, rule["ref"])
    if rule["method"].startswith("approx"): # approx:THRESHOLD
        threshold: float = float(rule["method"].split(":")[1])
        logger.debug("Approx: TH%f, REF%f, VAL%s", threshold, rule["ref"], repr(value))
        try:
            value = float(value)
        except (ValueError, TypeError):
            return False
        else:
            return abs(value - rule["ref"]) <= threshold
    if rule["method"] == "spreadsheet_range":
        subset_limit = MultiCellRange(rule["ref"][0])
        superset_limit = MultiCellRange(rule["ref"][1])
        return _multicellrange_containsby(subset_limit, value) \
            and _multicellrange_containsby(value, superset_limit)
    if rule["method"].startswith("range."):  # e.g., range.te [0, 2] -> 0 < x <= 2
        left_et = rule["method"][6]
        right_et = rule["method"][7]
        return getattr(operator, "l" + left_et)(rule["ref"][0], value) \
            and getattr(operator, "l" + right_et)(value, rule["ref"][1])
    if rule["method"] in {"str_list_eq", "str_set_eq"}:
        container_type_str: str = rule["method"][4:-3]
        container_type = getattr(builtins, container_type_str)

        value: container_type = container_type(value.strip("\"'").split(","))
        ref: container_type = container_type(rule["ref"])
        return value == ref
    raise NotImplementedError()


def are_lists_equal(list1, list2, comparison_func):
    # First check if both lists have the same length
    if len(list1) != len(list2):
        return False

    # Now make sure each element in one list has an equal element in the other list
    for item1 in list1:
        # Use the supplied function to test for an equal item
        if not any(comparison_func(item1, item2) for item2 in list2):
            return False

    # If all items match, the lists are equal
    return True


def compare_urls(url1, url2):
    if url1 is None or url2 is None:
        return url1 == url2

    def normalize_url(url):
        # Parse the URL
        parsed_url = urlparse(url)

        # If no scheme is present, assume 'http'
        scheme = parsed_url.scheme if parsed_url.scheme else 'http'

        # Lowercase the scheme and netloc, remove 'www.', and handle trailing slash
        normalized_netloc = parsed_url.netloc.lower().replace("www.", "")
        normalized_path = parsed_url.path if parsed_url.path != '/' else ''

        # Reassemble the URL with normalized components
        normalized_parsed_url = parsed_url._replace(scheme=scheme.lower(), netloc=normalized_netloc,
                                                    path=normalized_path)
        normalized_url = urlunparse(normalized_parsed_url)

        return normalized_url

    # Normalize both URLs for comparison
    norm_url1 = normalize_url(url1)
    norm_url2 = normalize_url(url2)

    # Compare the normalized URLs
    return norm_url1 == norm_url2
